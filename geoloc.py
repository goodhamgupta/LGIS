#Sample program to using geolocate API and json library to parse
import json
import urllib
import webbrowser
from kmlconverter import *
from openpyxl import *
#End of random import statements

def error_check(result):
	if( result["status"] == "ZERO_RESULTS") :
		print ("Query Successful but no results to show! Try to refine your search")
		return False
	elif (result["status"] == "OVER_QUERY_LIMIT") :
		print ("ERROR! Query Limit Exceeded!")
		return False
	elif ( result["status"] == "REQUEST_DENIED") :
		print ("ERROR! Invalid API key")
		return False
	elif( result["status"] == "INVALID_REQUEST") :
		print ("ERROR! Missing data!")
		return False
	elif (result["status"] == "UNKNOWN_ERROR") :
		print ("Server Error! Please try again later")
		return False
	elif ( result["status"] == "OK"):
		print "Status : OK"
		return True
def workbook_create():
	wb = Workbook()
	ws = wb.active
	ws['A1'] = "Latitude"
	ws['B1'] = "Longitude"
	ws['C1'] = "Name"
	ws['D1'] = "Rating"
	ws['E1'] = "Icon Type"
	wb.save('data.xlsx')


def geoloc_func():
	serviceurl = "https://maps.googleapis.com/maps/api/geocode/json?"
	loc = raw_input("Enter location: ")
	locurl = serviceurl + urllib.urlencode({'sensor': 'false' , 'address': loc}) + ",+CA&key=AIzaSyDUWKArBtvnMYthvAe1AmxfY6-cloBanwc"
	data = urllib.urlopen(locurl).read()
	#'Data' contains the JSON data retreived using the Google API
	#Loading the JSON data
	result = json.loads(data)
	#Error Checking
	fi = error_check(result)
	if (fi == True) :
		#Final Output
		lat = result["results"][0]["geometry"]["location"]["lat"]
		lng = result["results"][0]["geometry"]["location"]["lng"]
		fadd = result["results"][0]["formatted_address"]
		val = result["results"][0]["partial_match"]
		if(val == True):
			print "No exact match found. Following results are an APPROXIMATE!"
			print "Formatted Address: ", fadd
			print "Latitude:  ",lat
			print "Longitude: ",lng
			#Code to display the output on the browser
			
			print ("Showing final output on the browser!")
			final = "https://maps.google.com/?q=" + str(lat) + "," + str(lng)
			#webbrowser.open(final)
			
			#Getting nearby places using the above obtained lat and lng
			print "Nearby resturants are!\n"
			nearby_url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json?location="+ str(lat) + "," + str(lng) + "&radius=50000&types=food|restaurant&key=AIzaSyDUWKArBtvnMYthvAe1AmxfY6-cloBanwc"
			nearby = urllib.urlopen(nearby_url).read()
			nearby_data = json.loads(nearby)
			test = error_check(nearby_data)
			#Creating workbook to store data
			workbook_create()
			#Workbook with one empty Sheet has been created
			count = 2;
			if(test == True) :
				for item in nearby_data["results"]:
					lat = item["geometry"]["location"]["lat"]
					lng = item["geometry"]["location"]["lng"]
					name = item['name']
					rating = item["rating"]
					wb = load_workbook('data.xlsx')
					ws = wb.active
					ws['A' + str(count) ] = lat
					ws['B' + str(count) ] = lng
					ws['C' + str(count) ] = name
					ws['D' + str(count) ] = rating
					ws['E' + str(count) ] = "url-to-icon"
					count = count + 1
					# Rows can also be appended
					wb.save("data.xlsx")
				print "Data has been created and saved!"	
			#Excel sheet with the required data has been created
			#Test function to convert XLSX to KML file
			convert_file()
			print "Conversion Successful"

if __name__ == "__main__":
	geoloc_func()